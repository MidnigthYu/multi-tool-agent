"""Multi-format document parser + semantic chunker for RAG knowledge base.

Supports PDF, DOCX, XLSX plain-text extraction through a unified entry point.
Chunking uses paragraph-boundary-first splitting with configurable overlap.
All exceptions are caught internally — callers receive (False, error_msg) tuples.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from config.error_codes import ErrorCode
from config.settings import get_settings

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse_document(file_path: str) -> tuple[bool, str]:
    """Unified document parsing entry point.

    Dispatches to the appropriate parser based on file extension and returns
    extracted plain text.  All parser-level exceptions are caught and returned
    as (False, error_message) — this function never raises.

    Args:
        file_path: Absolute or relative path to a .pdf, .docx, or .xlsx file.

    Returns:
        (True, extracted_text) on success.
        (False, error_message) when the file is missing, unsupported, or
        parsing fails.
    """
    path = Path(file_path)
    if not path.exists():
        msg = f"[E0301] 文件不存在: {path.name}"
        logger.warning("[doc_parser] %s", msg)
        return False, msg

    ext = path.suffix.lower()
    if ext not in _PARSERS:
        msg = f"[E0301] 不支持的文件格式 '{ext}'，仅支持 PDF / DOCX / XLSX"
        logger.warning("[doc_parser] %s | file=%s", msg, path.name)
        return False, msg

    try:
        text = _PARSERS[ext](str(path))
    except Exception as exc:
        err_code = _EXT_ERROR_MAP.get(ext, ErrorCode.E0301)
        msg = f"[{err_code.value}] 文档解析失败: {path.name} — {exc}"
        logger.warning("[doc_parser] Parse failed | file=%s | exc=%s", path.name, exc)
        return False, msg

    if not text or not text.strip():
        msg = f"[E0307] 文件内容为空或无可提取文本: {path.name}"
        logger.warning("[doc_parser] %s", msg)
        return False, msg

    return True, text


def split_text_into_chunks(
    text: str,
    chunk_size: int | None = None,
    overlap: int | None = None,
) -> list[str]:
    """Split long text into semantic chunks for embedding.

    Strategy (in priority order):
    1. Split on paragraph boundaries (double-newline).
    2. If a paragraph still exceeds *chunk_size*, fall back to fixed-width
       sliding-window splitting with *overlap* characters between windows.

    Args:
        text: Raw document text.
        chunk_size: Maximum characters per chunk.  Defaults to
            ``settings.DOC_CHUNK_SIZE`` (1000).
        overlap: Overlap characters between adjacent chunks.  Defaults to
            ``settings.DOC_CHUNK_OVERLAP`` (200).

    Returns:
        List of chunk strings.  Empty list when *text* is empty.
    """
    if not text or not text.strip():
        return []

    settings = get_settings()
    cs = chunk_size if chunk_size is not None else settings.DOC_CHUNK_SIZE
    ol = overlap if overlap is not None else settings.DOC_CHUNK_OVERLAP

    # --- Phase 1: paragraph splitting ---
    paragraphs = _split_paragraphs(text)

    # --- Phase 2: enforce chunk_size ceiling ---
    chunks: list[str] = []
    for para in paragraphs:
        if len(para) <= cs:
            if para.strip():
                chunks.append(para.strip())
        else:
            chunks.extend(_fixed_width_chunks(para, cs, ol))

    return chunks


# ---------------------------------------------------------------------------
# Internal parsers (one per format)
# ---------------------------------------------------------------------------


def _parse_pdf(file_path: str) -> str:
    """Extract text from a PDF file via pypdf.PdfReader."""
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    pages: list[str] = []
    for page in reader.pages:
        extracted = page.extract_text()
        if extracted:
            pages.append(_sanitize_whitespace(extracted))
    return "\n\n".join(pages)


def _parse_docx(file_path: str) -> str:
    """Extract text from a DOCX file via python-docx."""
    from docx import Document

    doc = Document(file_path)
    paragraphs: list[str] = []
    for para in doc.paragraphs:
        text = para.text
        if text and text.strip():
            paragraphs.append(text.strip())
    return "\n\n".join(paragraphs)


def _parse_xlsx(file_path: str) -> str:
    """Extract text from an XLSX file via openpyxl in TSV-like format."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets_output: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = [f"[Sheet: {sheet_name}]"]
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
        sheets_output.append("\n".join(rows))
    wb.close()
    return "\n\n".join(sheets_output)


# ---------------------------------------------------------------------------
# Chunking helpers
# ---------------------------------------------------------------------------


def _split_paragraphs(text: str) -> list[str]:
    """Split *text* on double-newline boundaries, preserving multi-paragraph blocks."""
    return re.split(r"\n\s*\n", text)


def _fixed_width_chunks(text: str, chunk_size: int, overlap: int) -> list[str]:
    """Fallback: split *text* into fixed-width overlapping windows."""
    if len(text) <= chunk_size:
        return [text.strip()] if text.strip() else []

    chunks: list[str] = []
    step = max(1, chunk_size - overlap)
    start = 0
    while start < len(text):
        end = min(start + chunk_size, len(text))
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start += step
    return chunks


def _sanitize_whitespace(text: str) -> str:
    """Collapse runs of whitespace while preserving paragraph breaks."""
    # Normalise line-endings, collapse multiple blanks into single space
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[^\S\n]+", " ", text)  # horizontal whitespace → single space
    text = re.sub(r"\n{3,}", "\n\n", text)  # max 2 consecutive newlines
    return text.strip()


# ---------------------------------------------------------------------------
# Format dispatch table
# ---------------------------------------------------------------------------

_PARSERS: dict[str, Any] = {
    ".pdf": _parse_pdf,
    ".docx": _parse_docx,
    ".xlsx": _parse_xlsx,
}

_EXT_ERROR_MAP: dict[str, ErrorCode] = {
    ".pdf": ErrorCode.E0303,
    ".docx": ErrorCode.E0304,
    ".xlsx": ErrorCode.E0305,
}
